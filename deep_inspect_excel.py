import openpyxl

file_path = '周报2026-02-10.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb.active

print("=== detailed inspection of odin-search ===")

# Find odin-search again
for r in range(1, 40):
    for c in range(1, 20):
        val = ws.cell(row=r, column=c).value
        if str(val).strip() == 'odin-search':
            print(f"\nFound 'odin-search' anchor at Row {r}, Col {c} ({openpyxl.utils.get_column_letter(c)})")

            # Print the header row (r)
            print(f"Header Row ({r}):")
            for i in range(8):
                cell_val = ws.cell(row=r, column=c+i).value
                print(f"  Col {c+i}: {cell_val}")

            # Print the data rows (r+1 to r+4)
            print(f"\nData Rows ({r+1} to {r+4}):")
            for dr in range(r+1, r+5):
                row_vals = []
                for i in range(8):
                    val = ws.cell(row=dr, column=c+i).value
                    row_vals.append(str(val))
                print(f"  Row {dr}: {row_vals}")

print("\n=== detailed inspection of odin (reference) ===")
# Find odin
for r in range(1, 40):
    for c in range(1, 20):
        val = ws.cell(row=r, column=c).value
        if str(val).strip() == 'odin':
            print(f"\nFound 'odin' anchor at Row {r}, Col {c} ({openpyxl.utils.get_column_letter(c)})")
             # Print the data rows (r+1 to r+4)
            print(f"Data Rows ({r+1} to {r+4}):")
            for dr in range(r+1, r+5):
                row_vals = []
                for i in range(8):
                    val = ws.cell(row=dr, column=c+i).value
                    row_vals.append(str(val))
                print(f"  Row {dr}: {row_vals}")
            break
