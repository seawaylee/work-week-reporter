import openpyxl

file_path = '周报2026-02-10.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb.active

print("Inspecting odin-search block...")

# Find odin-search
found = False
for r in range(1, 40):
    for c in range(1, 20):
        val = ws.cell(row=r, column=c).value
        if str(val).strip() == 'odin-search':
            print(f"Found 'odin-search' at Row {r}, Col {c}")

            # Check Header
            qps_header = ws.cell(row=r, column=c+1).value
            print(f"Header at Col {c+1}: {qps_header}")

            # Check Data Rows (next 4 rows)
            for i in range(1, 5):
                curr_r = r + i
                date_val = ws.cell(row=curr_r, column=c).value
                qps_val = ws.cell(row=curr_r, column=c+1).value
                print(f"  Row {curr_r}: Date={date_val}, QPS={qps_val} (Type: {type(qps_val)})")

            found = True
            break
    if found: break

if not found:
    print("Could not find 'odin-search' in Excel.")
