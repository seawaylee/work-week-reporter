import openpyxl

file_path = '周报2026-02-10.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb.active

def check_service(name):
    print(f"Checking {name}...")
    for r in range(1, 40):
        for c in range(1, 20):
            val = ws.cell(row=r, column=c).value
            if str(val).strip() == name:
                # Data is in next 4 rows
                for i in range(1, 5):
                    dr = r + i
                    date_val = ws.cell(row=dr, column=c).value
                    req_val = ws.cell(row=dr, column=c+4).value # Col+4 is Requests
                    print(f"  Row {dr}: Date={date_val}, Reqs={req_val} (Type: {type(req_val)})")
                return
    print(f"  Not found.")

check_service('odin')
check_service('odin-search')
