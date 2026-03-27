import pandas as pd
import requests
import json
import openpyxl
import os
import sys
import re
import pathlib
import shlex
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from copy import copy
from datetime import datetime, timedelta

import urllib.parse

# --- Configuration ---
TXT_URL = os.getenv("TXT_URL", "")
TXT_HEADERS = {
    'Accept': 'application/json, text/plain, */*',
    'Content-Type': 'application/x-www-form-urlencoded',
    'Cookie': os.getenv("TXT_COOKIE", ""),
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/144.0.0.0 Safari/537.36'
}
TXT_BODY = 'board_id=20133&dataset_id=20609&cfg=%7B%22rows%22%3A%5B%7B%22columnName%22%3A%22from_dt%22%2C%22link%22%3Afalse%2C%22filterType%22%3A%22eq%22%2C%22values%22%3A%5B%5D%2C%22id%22%3A%225c12004a-2e41-4ca8-b221-703b61ac8337%22%2C%22alias%22%3A%22%E5%BC%80%E5%A7%8B%E6%97%A5%E6%9C%9F%22%7D%2C%7B%22columnName%22%3A%22to_dt%22%2C%22link%22%3Afalse%2C%22filterType%22%3A%22eq%22%2C%22values%22%3A%5B%5D%2C%22id%22%3A%22476cab63-266d-4d6b-b5b4-9655e655affc%22%2C%22alias%22%3A%22%E7%BB%93%E6%9D%9F%E6%97%A5%E6%9C%9F%22%7D%2C%7B%22columnName%22%3A%22profiletype%22%2C%22link%22%3Afalse%2C%22filterType%22%3A%22eq%22%2C%22values%22%3A%5B%5D%2C%22id%22%3A%225ec46a09-f07a-4f93-90d9-d9340a3ffcd8%22%2C%22alias%22%3A%22%E6%9C%8D%E5%8A%A1%E9%83%A8%E7%BD%B2%E7%B1%BB%E5%9E%8B%22%7D%2C%7B%22columnName%22%3A%22days%22%2C%22link%22%3Afalse%2C%22filterType%22%3A%22eq%22%2C%22values%22%3A%5B%5D%2C%22id%22%3A%22e8a15fbc-05d4-444c-b414-44e18f63c7fd%22%2C%22alias%22%3A%22%E5%A4%A9%E6%95%B0%22%7D%5D%2C%22columns%22%3A%5B%5D%2C%22filters%22%3A%5B%7B%22columnName%22%3A%22profiletype%22%2C%22filterType%22%3A%22%3D%22%2C%22values%22%3A%5B%5D%2C%22alias%22%3A%22profiletype%22%7D%5D%2C%22datalength%22%3A2000%2C%22values%22%3A%5B%7B%22column%22%3A%22median_sum%22%2C%22aggType%22%3A%22sum%22%2C%22alias%22%3A%2250%E5%88%86%E4%BD%8D%E5%93%8D%E5%BA%94%E6%97%B6%E9%97%B4%28ms%29%22%7D%2C%7B%22column%22%3A%22days%22%2C%22aggType%22%3A%22sum%22%2C%22alias%22%3A%2250%E5%88%86%E4%BD%8D%E5%93%8D%E5%BA%94%E6%97%B6%E9%97%B4%28ms%29%22%7D%2C%7B%22column%22%3A%22ninty_nine_sum%22%2C%22aggType%22%3A%22sum%22%2C%22alias%22%3A%2299%E5%88%86%E4%BD%8D%E5%93%8D%E5%BA%94%E6%97%B6%E9%97%B4%28ms%29%22%7D%2C%7B%22column%22%3A%22days%22%2C%22aggType%22%3A%22sum%22%2C%22alias%22%3A%2299%E5%88%86%E4%BD%8D%E5%93%8D%E5%BA%94%E6%97%B6%E9%97%B4%28ms%29%22%7D%2C%7B%22column%22%3A%22nine_nine_nine_sum%22%2C%22aggType%22%3A%22sum%22%2C%22alias%22%3A%2299.9%E5%88%86%E4%BD%8D%E5%93%8D%E5%BA%94%E6%97%B6%E9%97%B4%28ms%29%22%7D%2C%7B%22column%22%3A%22days%22%2C%22aggType%22%3A%22sum%22%2C%22alias%22%3A%2299.9%E5%88%86%E4%BD%8D%E5%93%8D%E5%BA%94%E6%97%B6%E9%97%B4%28ms%29%22%7D%2C%7B%22column%22%3A%22count_sum%22%2C%22aggType%22%3A%22sum%22%2C%22alias%22%3A%22%E5%B9%B3%E5%9D%87%E8%AF%B7%E6%B1%82%E6%95%B0%22%7D%2C%7B%22column%22%3A%22days%22%2C%22aggType%22%3A%22sum%22%2C%22alias%22%3A%22%E5%B9%B3%E5%9D%87%E8%AF%B7%E6%B1%82%E6%95%B0%22%7D%5D%7D&reload=true'

GRAFANA_URL = os.getenv("GRAFANA_URL", "https://grafana-m0ymy2z9.grafana.tencent-cloud.com/api/datasources/proxy/1/api/v1/query_range")
ENV_FILE = os.getenv("ENV_FILE", ".env.local")

# Optional: provide a custom encoded query template via env var for odin-interface.
ODIN_GRAFANA_QUERY_TEMPLATE = os.getenv("ODIN_GRAFANA_QUERY_TEMPLATE", "")

INPUT_FILE = '周报2026-01-09.xlsx'
OUTPUT_FILE = f'周报{datetime.now().strftime("%Y-%m-%d")}.xlsx'

# Updated Mapping per User Instructions
SERVICE_MAPPING = {
    'odin': '3202',
    'odin-home': '3201',
    'odin-search': '3206',
    'odin-video': '3205',
    'odin-article': '3203',
    'odin-focus': '3204',
    'odin-author': '3207',
    '视频Loki': '2',
    '视频重点场景Loki': '21',
    '频道Loki': '1',
    '话题Loki': '4',
    'algo-loki': '8',
    '社区Loki': '3',
    '焦点Loki': '6',
    'fis-Loki': '7', # User said fis-loki
}

GRAFANA_MAPPING = {
    'odin': 'umab-odin-interface',
    'odin-home': 'umab-odin-home-interface',
    'odin-search': 'umab-odin-search-interface',
    'odin-video': 'umab-odin-video-interface',
    'odin-article': 'umab-odin-article-interface',
    'odin-focus': 'umab-odin-focus-interface',
    'odin-author': 'umab-odin-author-interface'
}
GRAFANA_SERVICE_NAMES = {name.lower() for name in GRAFANA_MAPPING}

ANOMALY_FILL = PatternFill(start_color="FFF4CCCC", end_color="FFF4CCCC", fill_type="solid")
WEEK_RANGE_PATTERN = re.compile(r"^\d{4}-\d{4}$")
REPORT_FILE_PATTERN = re.compile(r"^周报(\d{4}-\d{2}-\d{2})\.xlsx$")


def load_env_values(env_file):
    env_path = pathlib.Path(env_file).expanduser()
    if not env_path.exists():
        return {}

    values = {}
    for raw_line in env_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, raw_value = line.split("=", 1)
        values[key.strip()] = " ".join(shlex.split(raw_value, posix=True, comments=False))
    return values


def get_grafana_request_config(env_file=ENV_FILE):
    env_values = load_env_values(env_file)
    return (
        env_values.get("GRAFANA_URL") or os.getenv("GRAFANA_URL", GRAFANA_URL),
        {
            'accept': 'application/json, text/plain, */*',
            'content-type': 'application/x-www-form-urlencoded',
            'cookie': env_values.get("GRAFANA_COOKIE", os.getenv("GRAFANA_COOKIE", "")),
            'x-grafana-org-id': env_values.get("GRAFANA_ORG_ID", os.getenv("GRAFANA_ORG_ID", "1"))
        },
    )

def fetch_txt_data():
    """从天象台 API 获取数据"""
    print("Fetching Tianxiangtai data...")
    response = requests.post(TXT_URL, headers=TXT_HEADERS, data=TXT_BODY, verify=False)
    if response.status_code != 200:
        raise Exception(f"TXT API Failed: {response.text}")
    return response.json()

def get_grafana_body(app_name, start_ts, end_ts):
    """Construct Grafana query body"""
    if app_name == 'umab-odin-interface' and ODIN_GRAFANA_QUERY_TEMPLATE:
        # Use optional custom query template for odin.
        return ODIN_GRAFANA_QUERY_TEMPLATE.format(start_ts=start_ts, end_ts=end_ts)
    else:
        # Use generic query for others
        # Query: max_over_time(sum(rate(http_server_requests_seconds_count{application="APP_NAME"}[1m]))[24h:])
        query = f'max_over_time(sum(rate(http_server_requests_seconds_count{{application="{app_name}"}}[1m]))[24h:])'
        encoded_query = urllib.parse.quote(query)
        return f'query={encoded_query}&start={start_ts}&end={end_ts}&step=300'

def fetch_grafana_data(app_name, start_ts, end_ts, env_file=ENV_FILE):
    """从 Grafana API 获取数据 for specific app"""
    # print(f"Fetching Grafana data for {app_name}...")
    body = get_grafana_body(app_name, start_ts, end_ts)
    grafana_url, grafana_headers = get_grafana_request_config(env_file=env_file)
    response = requests.post(grafana_url, headers=grafana_headers, data=body, verify=False)
    if response.status_code != 200:
        print(f"Grafana API Failed for {app_name}: {response.text}")
        try:
            data = response.json()
        except Exception:
            data = {"message": response.text}
        data["_http_status"] = response.status_code
        return data
    try:
        return response.json()
    except Exception:
        return {"message": "invalid_json_response", "_http_status": response.status_code}


def parse_grafana_max_qps(raw):
    """Extract max QPS from Grafana response."""
    if not isinstance(raw, dict):
        return None, "invalid_response"

    status_code = raw.get("_http_status")
    message = str(raw.get("message", "")).strip().lower()
    if status_code == 401 or "unauthorized" in message:
        return None, "unauthorized"

    series = raw.get("data", {}).get("result", [])
    if not series:
        return None, "no_data"

    values = series[0].get("values", [])
    if not values:
        return None, "no_data"

    try:
        max_qps_val = max(float(x[1]) for x in values)
    except Exception:
        return None, "parse_error"

    return max_qps_val, None


def write_qps_value(ws, row, col, qps_val):
    """Always write QPS cell to avoid stale carried-over values."""
    ws.cell(row=row, column=col).value = int(qps_val) if qps_val is not None else None


def resolve_primary_metric_value(service_name, metrics, grafana_qps_map):
    """Resolve top-block second-column metric: QPS for Odin, tp50 for Loki."""
    service_key = str(service_name).strip().lower()

    if service_key in GRAFANA_SERVICE_NAMES:
        qps_val = grafana_qps_map.get(service_name)
        if qps_val is None:
            lower_map = {str(k).lower(): v for k, v in grafana_qps_map.items()}
            qps_val = lower_map.get(service_key)
        if qps_val is None:
            return None, "grafana_qps_missing"
        return int(qps_val), "grafana_qps"

    p50_val = to_number(metrics.get("p50") if isinstance(metrics, dict) else None)
    if p50_val is None:
        return None, "p50_missing"
    return int(p50_val), "p50"


def to_number(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if not s:
        return None
    s = s.replace(",", "")
    try:
        return float(s)
    except Exception:
        return None


def is_large_change(current_value, previous_value, threshold=0.5):
    """Return True when relative change is strictly greater than threshold."""
    current_num = to_number(current_value)
    previous_num = to_number(previous_value)
    if current_num is None or previous_num is None or previous_num == 0:
        return False
    change = abs(current_num - previous_num) / abs(previous_num)
    return change > threshold


def apply_anomaly_fill(cell, current_value, previous_value):
    if is_large_change(current_value, previous_value):
        cell.fill = copy(ANOMALY_FILL)
    else:
        cell.fill = PatternFill(fill_type=None)


def clear_annotation_cell(ws, row, col_start):
    # Column +5 is the note/reason column; always clear for newly written week.
    cell = ws.cell(row=row, column=col_start + 5)
    cell.value = None
    cell.comment = None

def parse_txt_data(raw_data):
    """解析天象台数据为字典 {ProfileID: {DateStr: Metrics}}"""
    data = {}
    rows = raw_data.get('body', {}).get('data', [])

    # Find the latest available date across all data
    latest_date_str = "00000000"

    for row in rows:
        from_key = next((k for k in row.keys() if k.startswith("from_dt")), None)
        if from_key:
            dt = str(row[from_key])
            if dt > latest_date_str:
                latest_date_str = dt

    print(f"Detected Latest Date in API: {latest_date_str}")

    for row in rows:
        pid_key = next((k for k in row.keys() if k.startswith("profiletype")), None)
        from_key = next((k for k in row.keys() if k.startswith("from_dt")), None)
        to_key = next((k for k in row.keys() if k.startswith("to_dt")), None)

        if not pid_key or not from_key: continue

        pid = str(row[pid_key])
        from_dt = str(row[from_key])
        to_dt = str(row[to_key]) if to_key else ""

        req_key = next((k for k in row.keys() if k.startswith("count_sum")), None)
        p99_sum_key = next((k for k in row.keys() if k.startswith("ninty_nine_sum")), None)
        p999_sum_key = next((k for k in row.keys() if k.startswith("nine_nine_nine_sum")), None)
        p50_sum_key = next((k for k in row.keys() if k.startswith("median_sum")), None)
        days_key = next((k for k in row.keys() if k.startswith("days") and k != "days_sum"), None)

        days = 1
        if days_key and row.get(days_key):
             days = int(row[days_key])

        reqs = int(row.get(req_key, 0)) if req_key else 0
        p99_sum = float(row.get(p99_sum_key, 0) or 0)
        p999_sum = float(row.get(p999_sum_key, 0) or 0)
        p50_sum = float(row.get(p50_sum_key, 0) or 0)

        # Analysis of Excel History vs API:
        # odin-search Week 3: 68M (Excel) vs ~4.8亿 (API Total for 7 days) -> 4.8亿/7 = 68M
        # odin-search Week 4: 2.7亿 (API Total for 4 days) -> 2.7亿/4 = 68.5M
        # Conclusion: Excel strictly uses Daily Average.
        avg_reqs = int(reqs / days) if days else 0

        if pid not in data: data[pid] = {}
        data[pid][from_dt] = {
            "requests": avg_reqs, # FORCE Daily Average to match Excel magnitude
            "p50": p50_sum / days if days else 0,
            "p99": p99_sum / days if days else 0,
            "p999": p999_sum / days if days else 0,
            "to_dt": to_dt,
            "days": days
        }

    return data, latest_date_str

def find_service_blocks(ws):
    """识别 Excel 中的服务区块"""
    blocks = []
    # 扫描合理的表头范围
    for r in range(1, 40):
        for c in range(1, 20):
            val = ws.cell(row=r, column=c).value
            next_val = ws.cell(row=r, column=c+1).value
            if next_val and (("QPS" in str(next_val)) or ("tp50" in str(next_val))):
                service_name = str(val).strip()
                blocks.append({
                    'name': service_name,
                    'row': r,
                    'col': c
                })
    return blocks

def update_bottom_raw_data(ws, txt_data, latest_api_date):
    """Update the bottom raw data section (starts around row 33)"""
    print("Updating Bottom Raw Data Section...")
    for r in range(33, 100):
        # Column 2 (C) seems to be Profile ID (e.g., 1, 2, 3...)
        pid_cell = ws.cell(row=r, column=3).value
        if pid_cell is None:
            break

        pid = str(pid_cell)

        # Get new data for this PID
        if pid in txt_data and latest_api_date in txt_data[pid]:
            metrics = txt_data[pid][latest_api_date]

            # Column A (1): Start Date
            ws.cell(row=r, column=1).value = latest_api_date
            # Column B (2): End Date
            ws.cell(row=r, column=2).value = metrics.get('to_dt', '')

            # Column D (4): Days
            ws.cell(row=r, column=4).value = metrics.get('days', 7)

            # Column F (6): TP99 (p99) -> INT
            ws.cell(row=r, column=6).value = int(metrics['p99'])
            # Column G (7): TP999 (p999) -> INT
            ws.cell(row=r, column=7).value = int(metrics['p999'])
            # Column H (8): Requests
            ws.cell(row=r, column=8).value = metrics['requests']

def copy_style(src_cell, dst_cell):
    """Copy cell style from source to destination"""
    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.border = copy(src_cell.border)
        dst_cell.fill = copy(src_cell.fill)
        # We don't copy number format anymore because user wants specific format (integers)
        # dst_cell.number_format = copy(src_cell.number_format)
        dst_cell.protection = copy(src_cell.protection)
        dst_cell.alignment = copy(src_cell.alignment)


def copy_cell_state(src_cell, dst_cell):
    """Copy full visible cell state for historical row rolling."""
    dst_cell.value = src_cell.value
    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.border = copy(src_cell.border)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.protection = copy(src_cell.protection)
        dst_cell.alignment = copy(src_cell.alignment)
        dst_cell.number_format = copy(src_cell.number_format)
    dst_cell.comment = None
    dst_cell.comment = copy(src_cell.comment) if src_cell.comment else None


def collect_block_data_rows(ws, r_start, c_start, scan_limit=10):
    """Collect contiguous data rows for one service block by row position only."""
    data_rows = []
    started = False
    for i in range(scan_limit):
        curr_r = r_start + i
        val = ws.cell(row=curr_r, column=c_start).value
        row_label = str(val).strip() if isinstance(val, str) else ""
        is_data_row = bool(WEEK_RANGE_PATTERN.match(row_label))
        if is_data_row:
            data_rows.append(curr_r)
            started = True
            continue

        # Once block data starts, a non-data row means next block or end.
        if started:
            break

        if val is None and i > 0:
            break

    return data_rows


def prepare_block_target_row(ws, r_start, c_start, data_col_count=6, min_rows_for_roll=4, scan_limit=10):
    """
    If rows < min_rows_for_roll, append at the next row.
    If rows >= min_rows_for_roll, delete first row by shifting all following rows up.
    Returns the row index where latest data should be written.
    """
    data_rows = collect_block_data_rows(ws, r_start, c_start, scan_limit=scan_limit)
    if len(data_rows) < min_rows_for_roll:
        return data_rows[-1] + 1 if data_rows else r_start

    for idx in range(len(data_rows) - 1):
        src_r = data_rows[idx + 1]
        dst_r = data_rows[idx]
        for offset in range(data_col_count):
            copy_cell_state(
                ws.cell(row=src_r, column=c_start + offset),
                ws.cell(row=dst_r, column=c_start + offset),
            )

    return data_rows[-1]


def choose_input_report_file(base_dir=".", today=None, fallback_file=INPUT_FILE):
    """
    Choose the latest weekly report strictly before `today`.
    Fallback to `fallback_file` only when no prior weekly report exists.
    """
    base_path = pathlib.Path(base_dir)
    today_date = today or datetime.now().date()

    candidates = []
    for path in base_path.iterdir():
        if not path.is_file():
            continue
        match = REPORT_FILE_PATTERN.match(path.name)
        if not match:
            continue
        try:
            report_date = datetime.strptime(match.group(1), "%Y-%m-%d").date()
        except Exception:
            continue
        if report_date < today_date:
            candidates.append((report_date, path))

    if candidates:
        candidates.sort(key=lambda x: x[0])
        return str(candidates[-1][1])

    fallback_path = pathlib.Path(fallback_file)
    if not fallback_path.is_absolute():
        fallback_path = base_path / fallback_path
    if fallback_path.exists():
        return str(fallback_path)

    raise FileNotFoundError(
        f"No previous weekly report found before {today_date}, and fallback missing: {fallback_path}"
    )

def process_report():
    print("Starting Weekly Report Generation...")
    _, grafana_headers = get_grafana_request_config()

    required = {
        "TXT_URL": TXT_URL,
        "TXT_COOKIE": TXT_HEADERS.get("Cookie", ""),
        "GRAFANA_COOKIE": grafana_headers.get("cookie", "")
    }
    missing = [k for k, v in required.items() if not v]
    if missing:
        print(f"Missing required env vars: {', '.join(missing)}")
        print("Set env vars first, then rerun.")
        return 2

    # 1. 获取数据
    try:
        txt_raw = fetch_txt_data()
        txt_data, latest_api_date = parse_txt_data(txt_raw)
        print(f"Loaded {len(txt_data)} profiles. Latest Date: {latest_api_date}")
    except Exception as e:
        print(f"Error fetching TXT data: {e}")
        return 1

    # 2. 加载 Excel
    try:
        input_file = choose_input_report_file(base_dir=".", today=datetime.now().date(), fallback_file=INPUT_FILE)
        wb = openpyxl.load_workbook(input_file)
        ws = wb.active
        print(f"Loaded Excel: {input_file}")
    except Exception as e:
        print(f"Error loading Excel: {e}")
        return 1

    # 3. 构造显示用日期字符串
    sample_metric = next(iter(txt_data.values())).get(latest_api_date)
    to_dt = sample_metric.get('to_dt', '')

    # Format: "MMDD-MMDD"
    if len(latest_api_date) == 8 and len(to_dt) == 8:
        new_date_str = f"{latest_api_date[4:]}-{to_dt[4:]}"
    else:
        new_date_str = f"{latest_api_date}-Latest"

    print(f"New Row Date String: {new_date_str}")

    # 获取 Grafana QPS (Pre-fetch all services)
    grafana_qps_map = {}
    grafana_errors = {}
    try:
        dt_obj = datetime.strptime(latest_api_date, "%Y%m%d")
        end_dt_obj = datetime.strptime(to_dt, "%Y%m%d")
        ts_start = int(dt_obj.timestamp())
        ts_end = int(end_dt_obj.timestamp()) + 86399

        # Iterate over services that have Grafana mapping
        print("Fetching Grafana data for all mapped services...")
        for svc_name, app_name in GRAFANA_MAPPING.items():
             grafana_raw = fetch_grafana_data(app_name, ts_start, ts_end)
             max_qps_val, err = parse_grafana_max_qps(grafana_raw)
             if err:
                grafana_errors[svc_name] = err
                print(f"  {svc_name}: {err}")
             else:
                grafana_qps_map[svc_name] = max_qps_val
                print(f"  {svc_name}: {max_qps_val}")

    except Exception as e:
        print(f"Error fetching Grafana: {e}")

    # 4. 更新底部原始数据
    update_bottom_raw_data(ws, txt_data, latest_api_date)

    # 5. 更新顶部周报展示块
    blocks = find_service_blocks(ws)
    print(f"Found {len(blocks)} service blocks.")

    for block in blocks:
        name = block['name']
        # Try exact match, then case-insensitive if needed
        pid = SERVICE_MAPPING.get(name)

        if not pid:
             # Try case insensitive match for keys
             lower_map = {k.lower(): v for k, v in SERVICE_MAPPING.items()}
             pid = lower_map.get(name.lower())

        if not pid:
            print(f"Skipping {name}: No mapping ID found.")
            continue

        new_metrics = txt_data.get(pid, {}).get(latest_api_date)
        if not new_metrics:
            print(f"Skipping {name}: No data for date {latest_api_date}")
            continue

        # 定位写入行
        r_start = block['row'] + 1
        c_start = block['col']

        target_row_idx = prepare_block_target_row(
            ws,
            r_start=r_start,
            c_start=c_start,
            data_col_count=6,
            min_rows_for_roll=4,
            scan_limit=10,
        )

        # 写入新数据
        # 重要的是：把上一行的样式复制给新的一行（因为新的一行可能是空的或者样式不同）
        # 如果是滚动（len=4），target_row_idx 就是最后一行，它已经有样式了。
        # 如果是追加（len<4），target_row_idx 是新行，需要从上一行复制样式。

        prev_row_idx = target_row_idx - 1
        if prev_row_idx >= r_start:
             for offset in range(6):
                 src_cell = ws.cell(row=prev_row_idx, column=c_start+offset)
                 dst_cell = ws.cell(row=target_row_idx, column=c_start+offset)
                 copy_style(src_cell, dst_cell)

        # Update values
        ws.cell(row=target_row_idx, column=c_start).value = new_date_str

        primary_metric_val, primary_metric_source = resolve_primary_metric_value(name, new_metrics, grafana_qps_map)
        write_qps_value(ws, target_row_idx, c_start+1, primary_metric_val)
        prev_primary_metric = ws.cell(row=prev_row_idx, column=c_start+1).value if prev_row_idx >= r_start else None
        apply_anomaly_fill(
            ws.cell(row=target_row_idx, column=c_start+1),
            ws.cell(row=target_row_idx, column=c_start+1).value,
            prev_primary_metric,
        )
        if primary_metric_source == "grafana_qps_missing":
            print(f"Warning: missing QPS for {name}; cleared QPS cell.")
        elif primary_metric_source == "p50_missing":
            print(f"Warning: missing tp50 for {name}; cleared tp50 cell.")

        # 核心逻辑修改：强制取整
        p99_val = int(new_metrics['p99'])
        p999_val = int(new_metrics['p999'])
        req_val = new_metrics['requests']

        ws.cell(row=target_row_idx, column=c_start+2).value = p99_val
        ws.cell(row=target_row_idx, column=c_start+3).value = p999_val
        # Requests is already int
        ws.cell(row=target_row_idx, column=c_start+4).value = req_val

        prev_p99 = ws.cell(row=prev_row_idx, column=c_start+2).value if prev_row_idx >= r_start else None
        prev_p999 = ws.cell(row=prev_row_idx, column=c_start+3).value if prev_row_idx >= r_start else None
        prev_req = ws.cell(row=prev_row_idx, column=c_start+4).value if prev_row_idx >= r_start else None
        apply_anomaly_fill(ws.cell(row=target_row_idx, column=c_start+2), p99_val, prev_p99)
        apply_anomaly_fill(ws.cell(row=target_row_idx, column=c_start+3), p999_val, prev_p999)
        apply_anomaly_fill(ws.cell(row=target_row_idx, column=c_start+4), req_val, prev_req)

        clear_annotation_cell(ws, target_row_idx, c_start)

        print(f"Updated {name} (Profile {pid})")

    wb.save(OUTPUT_FILE)
    print(f"\nSuccess! Report generated: {OUTPUT_FILE}")
    if grafana_errors:
        missing = ", ".join(sorted(grafana_errors.keys()))
        print(f"Incomplete Grafana QPS data for: {missing}")
        print("Please refresh Grafana cookie and rerun.")
        return 4
    return 0

if __name__ == "__main__":
    raise SystemExit(process_report())
