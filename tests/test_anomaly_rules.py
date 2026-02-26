import pathlib
import sys
import unittest

import openpyxl
from openpyxl.styles import PatternFill

PROJECT_ROOT = pathlib.Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

import generate_weekly_report as gwr


class AnomalyRulesTests(unittest.TestCase):
    def test_is_large_change_true_when_over_50_percent(self):
        self.assertTrue(gwr.is_large_change(151, 100))
        self.assertTrue(gwr.is_large_change(49, 100))

    def test_is_large_change_false_when_equal_or_below_50_percent(self):
        self.assertFalse(gwr.is_large_change(150, 100))
        self.assertFalse(gwr.is_large_change(50, 100))
        self.assertFalse(gwr.is_large_change(120, 100))

    def test_apply_anomaly_fill_sets_and_clears_fill(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        cell = ws.cell(row=1, column=1)
        cell.fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

        gwr.apply_anomaly_fill(cell, 110, 100)
        self.assertIsNone(cell.fill.fill_type)

        gwr.apply_anomaly_fill(cell, 200, 100)
        self.assertEqual(cell.fill.fill_type, "solid")
        self.assertEqual(cell.fill.start_color.rgb, "FFF4CCCC")

    def test_clear_annotation_cell(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=4, column=6).value = "用户推荐0点爆量"
        gwr.clear_annotation_cell(ws, 4, 1)
        self.assertIsNone(ws.cell(row=4, column=6).value)


if __name__ == "__main__":
    unittest.main()
