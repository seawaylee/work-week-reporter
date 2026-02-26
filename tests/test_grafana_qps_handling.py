import pathlib
import sys
import unittest

import openpyxl

PROJECT_ROOT = pathlib.Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

import generate_weekly_report as gwr


class GrafanaQpsHandlingTests(unittest.TestCase):
    def test_parse_grafana_max_qps_unauthorized(self):
        raw = {"message": "Unauthorized"}
        qps, err = gwr.parse_grafana_max_qps(raw)
        self.assertIsNone(qps)
        self.assertEqual(err, "unauthorized")

    def test_parse_grafana_max_qps_success(self):
        raw = {
            "status": "success",
            "data": {
                "result": [
                    {"values": [[1, "12.5"], [2, "20"], [3, "19.9"]]}
                ]
            },
        }
        qps, err = gwr.parse_grafana_max_qps(raw)
        self.assertEqual(qps, 20.0)
        self.assertIsNone(err)

    def test_write_qps_value_clears_stale_cell(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1).value = 999
        gwr.write_qps_value(ws, 1, 1, None)
        self.assertIsNone(ws.cell(row=1, column=1).value)


if __name__ == "__main__":
    unittest.main()
