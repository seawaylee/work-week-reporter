import pathlib
import sys
import unittest

import openpyxl
from openpyxl.comments import Comment

PROJECT_ROOT = pathlib.Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

import generate_weekly_report as gwr


class BlockRollingTests(unittest.TestCase):
    def _new_ws(self):
        wb = openpyxl.Workbook()
        return wb.active

    def test_prepare_block_target_row_less_than_4_appends_without_deleting(self):
        ws = self._new_ws()
        ws.cell(row=1, column=1).value = "odin"
        ws.cell(row=1, column=2).value = "峰值QPS"

        ws.cell(row=2, column=1).value = "0206-0212"
        ws.cell(row=3, column=1).value = "0213-0219"
        ws.cell(row=4, column=1).value = "0220-0226"

        fn = getattr(gwr, "prepare_block_target_row", None)
        self.assertIsNotNone(fn)
        target_row = fn(ws, r_start=2, c_start=1, data_col_count=6)

        self.assertEqual(target_row, 5)
        self.assertEqual(ws.cell(row=2, column=1).value, "0206-0212")
        self.assertEqual(ws.cell(row=3, column=1).value, "0213-0219")
        self.assertEqual(ws.cell(row=4, column=1).value, "0220-0226")

    def test_prepare_block_target_row_four_rows_deletes_first_and_shifts_up(self):
        ws = self._new_ws()
        ws.cell(row=1, column=1).value = "odin"
        ws.cell(row=1, column=2).value = "峰值QPS"

        for row, label in [(2, "0206-0212"), (3, "0213-0219"), (4, "0220-0226"), (5, "0227-0305")]:
            ws.cell(row=row, column=1).value = label
            ws.cell(row=row, column=2).value = row * 100

        fn = getattr(gwr, "prepare_block_target_row", None)
        self.assertIsNotNone(fn)
        target_row = fn(ws, r_start=2, c_start=1, data_col_count=6)

        self.assertEqual(target_row, 5)
        self.assertEqual(ws.cell(row=2, column=1).value, "0213-0219")
        self.assertEqual(ws.cell(row=3, column=1).value, "0220-0226")
        self.assertEqual(ws.cell(row=4, column=1).value, "0227-0305")

    def test_prepare_block_target_row_preserves_note_text_and_comment_when_rolling(self):
        ws = self._new_ws()
        ws.cell(row=1, column=1).value = "odin"
        ws.cell(row=1, column=2).value = "峰值QPS"

        for row, label, note in [
            (2, "0206-0212", "note-a"),
            (3, "0213-0219", "note-b"),
            (4, "0220-0226", "note-c"),
            (5, "0227-0305", "note-d"),
        ]:
            ws.cell(row=row, column=1).value = label
            ws.cell(row=row, column=6).value = note
            ws.cell(row=row, column=6).comment = Comment(f"comment-{note}", "tester")

        fn = getattr(gwr, "prepare_block_target_row", None)
        self.assertIsNotNone(fn)
        target_row = fn(ws, r_start=2, c_start=1, data_col_count=6)

        self.assertEqual(target_row, 5)
        self.assertEqual(ws.cell(row=2, column=6).value, "note-b")
        self.assertEqual(ws.cell(row=3, column=6).value, "note-c")
        self.assertEqual(ws.cell(row=4, column=6).value, "note-d")
        self.assertEqual(ws.cell(row=2, column=6).comment.text, "comment-note-b")
        self.assertEqual(ws.cell(row=3, column=6).comment.text, "comment-note-c")
        self.assertEqual(ws.cell(row=4, column=6).comment.text, "comment-note-d")

    def test_prepare_block_target_row_append_does_not_copy_historical_comment_to_new_row(self):
        ws = self._new_ws()
        ws.cell(row=1, column=1).value = "odin"
        ws.cell(row=1, column=2).value = "峰值QPS"
        ws.cell(row=2, column=1).value = "0206-0212"
        ws.cell(row=3, column=1).value = "0213-0219"
        ws.cell(row=4, column=1).value = "0220-0226"
        ws.cell(row=4, column=6).comment = Comment("historical", "tester")

        target_row = gwr.prepare_block_target_row(ws, r_start=2, c_start=1, data_col_count=6)

        prev_row_idx = target_row - 1
        for offset in range(6):
            gwr.copy_style(ws.cell(row=prev_row_idx, column=1 + offset), ws.cell(row=target_row, column=1 + offset))

        self.assertEqual(target_row, 5)
        self.assertEqual(ws.cell(row=4, column=6).comment.text, "historical")
        self.assertIsNone(ws.cell(row=5, column=6).comment)

    def test_prepare_block_target_row_stops_at_next_block_header(self):
        ws = self._new_ws()
        ws.cell(row=1, column=1).value = "odin"
        ws.cell(row=1, column=2).value = "峰值QPS"

        ws.cell(row=2, column=1).value = "0206-0212"
        ws.cell(row=3, column=1).value = "0213-0219"
        ws.cell(row=4, column=1).value = "0220-0226"

        ws.cell(row=5, column=1).value = "odin-video"
        ws.cell(row=5, column=2).value = "峰值QPS"
        ws.cell(row=6, column=1).value = "0206-0212"
        ws.cell(row=7, column=1).value = "0213-0219"

        fn = getattr(gwr, "prepare_block_target_row", None)
        self.assertIsNotNone(fn)
        target_row = fn(ws, r_start=2, c_start=1, data_col_count=6)

        # The current block has only 3 rows; must append, not roll using next block's rows.
        self.assertEqual(target_row, 5)
        self.assertEqual(ws.cell(row=2, column=1).value, "0206-0212")
        self.assertEqual(ws.cell(row=3, column=1).value, "0213-0219")
        self.assertEqual(ws.cell(row=4, column=1).value, "0220-0226")


if __name__ == "__main__":
    unittest.main()
